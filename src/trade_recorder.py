"""
Trade Recorder — Excel-Based Trade Journal

Automatically records completed trades into an Excel workbook.
Creates the file if it doesn't exist, appends rows, and manages
column headers.

All file paths are parameterised — no hard-coded locations.

Usage:
    recorder = TradeRecorder("path/to/journal.xlsx", headers=[...])
    recorder.record({
        "date": "2025-09-03",
        "symbol": "SPX",
        "strategy": "Iron Condor",
        "premium": -2.50,
        "quantity": 1,
    })
"""

import os
from datetime import datetime
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter

logger = __import__("logging").getLogger(__name__)


class TradeRecorder:
    """
    Persist trade records to an Excel workbook.

    Parameters
    ----------
    file_path:
        Path to the Excel (.xlsx) file.  Created automatically if it
        does not exist.
    headers:
        Column headers for the worksheet.  The first row is used as
        the header row.  Missing columns are added automatically when
        an existing file is opened.
    sheet_name:
        Worksheet name (default: active sheet).
    """

    def __init__(
        self,
        file_path: str,
        headers: Optional[List[str]] = None,
        sheet_name: Optional[str] = None,
    ):
        self.file_path = file_path
        self.headers = headers or [
            "Date", "Time", "Symbol", "Strategy",
            "Action", "Strike", "Expiry",
            "Price", "Quantity", "Premium",
        ]
        self.sheet_name = sheet_name
        self._workbook: Optional[openpyxl.Workbook] = None
        self._sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

    # ------------------------------------------------------------------
    # Lazy initialisation
    # ------------------------------------------------------------------

    def _ensure_loaded(self) -> None:
        """Open or create the workbook (idempotent)."""
        if self._workbook is not None:
            return

        if not os.path.exists(self.file_path):
            self._workbook = openpyxl.Workbook()
            self._sheet = self._workbook.active
            if self.sheet_name:
                self._sheet.title = self.sheet_name
            self._sheet.append(self.headers)
            logger.info("Created new trade journal: %s", self.file_path)
        else:
            self._workbook = openpyxl.load_workbook(self.file_path)
            self._sheet = self._workbook.active
            if self.sheet_name and self.sheet_name != self._sheet.title:
                self._sheet.title = self.sheet_name
            self._sync_headers()

    def _sync_headers(self) -> None:
        """Ensure all expected columns exist (adds missing ones)."""
        existing = [c.value for c in self._sheet[1]] if self._sheet else []
        if not existing:
            self._sheet.append(self.headers)
            return

        missing = len(self.headers) - len(existing)
        if missing > 0:
            for i in range(missing):
                col = len(existing) + i + 1
                self._sheet[f"{get_column_letter(col)}1"] = self.headers[len(existing) + i]
            self._workbook.save(self.file_path)
            logger.info("Added %d missing column(s) to trade journal", missing)

    # ------------------------------------------------------------------
    # Recording
    # ------------------------------------------------------------------

    def record(self, data: Dict[str, Any]) -> None:
        """
        Append one trade record to the journal.

        The dict keys should align with ``self.headers``.  Extra keys
        are ignored; missing keys produce empty cells.

        Errors are logged but never propagated — a recording failure
        should never interrupt the trading flow.
        """
        try:
            self._ensure_loaded()

            row = [data.get(h, "") for h in self.headers]
            self._sheet.append(row)

            # Format date columns (indices 0 and 2 assumed date-like)
            last_row = self._sheet.max_row
            for col_idx in {0, 2}:
                if col_idx < len(self.headers):
                    cell = self._sheet.cell(row=last_row, column=col_idx + 1)
                    if isinstance(cell.value, datetime):
                        cell.number_format = "yyyy/mm/dd"

            self._workbook.save(self.file_path)
            logger.info("Trade recorded: %s", data.get("Strategy", data))

        except Exception as exc:
            logger.error("Failed to record trade: %s", exc)
            # Never crash the main process for a recording error

    # ------------------------------------------------------------------
    # Convenience
    # ------------------------------------------------------------------

    def record_trade(
        self,
        timestamp: Optional[datetime] = None,
        symbol: str = "",
        strategy: str = "",
        price: float = 0.0,
        quantity: int = 0,
        **extra,
    ) -> None:
        """Quick helper for standard trade data."""
        ts = timestamp or datetime.now()
        self.record({
            "Date": ts.strftime("%Y-%m-%d"),
            "Time": ts.strftime("%H:%M:%S"),
            "Symbol": symbol,
            "Strategy": strategy,
            "Price": price,
            "Quantity": quantity,
            **extra,
        })
